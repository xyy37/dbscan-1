
import os
import math
import time
import itertools
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from collections import deque
from image_preprocessing import compare_cell_formats

class DBSCAN:
    def __init__(self, eps, min_samples):
        self.eps = eps
        self.min_samples = min_samples

    def fit_predict(self, X):
        self.labels_ = [-1] * len(X)
        self.visited = set()

        cluster_id = 0
        for i in range(len(X)):
            if i in self.visited:
                continue
            self.visited.add(i)

            neighbors = self.region_query(X, i)
            if len(neighbors) < self.min_samples:
                self.labels_[i] = -1
            else:
                self.expand_cluster(X, i, neighbors, cluster_id)
                cluster_id += 1

        return self.labels_

    def region_query(self, X, i):
        neighbors = []
        for j in range(len(X)):
            if i != j:
                x_1, y_1 = X[i]
                x_2, y_2 = X[j]
                if steepness_matrix[x_1, y_1, x_2, y_2] < self.eps:
                    neighbors.append(j)
        return neighbors

    def expand_cluster(self, X, i, neighbors, cluster_id):
        self.labels_[i] = cluster_id
        j = 0
        while j < len(neighbors):
            neighbor = neighbors[j]
            if neighbor not in self.visited:
                self.visited.add(neighbor)
                new_neighbors = self.region_query(X, neighbor)
                if len(new_neighbors) >= self.min_samples:
                    neighbors += new_neighbors
            if self.labels_[neighbor] == -1:
                self.labels_[neighbor] = cluster_id
            j += 1

    def distance(self, a, b):
        return ((a[0] - b[0]) ** 2 + (a[1] - b[1]) ** 2) ** 0.5

def adjacency_list(original_plot_sheet):
    graph = {}
    rows, cols = original_plot_sheet.shape

    def is_adjacent(point1, point2):
        return abs(point1[0] - point2[0]) + abs(point1[1] - point2[1]) == 1

    for i in range(rows):
        for j in range(cols):
            current_point = (i, j)

            for neighbor_i, neighbor_j in [(i - 1, j), (i + 1, j), (i, j - 1), (i, j + 1)]:
                neighbor_point = (neighbor_i, neighbor_j)

                if 0 <= neighbor_i < rows and 0 <= neighbor_j < cols and is_adjacent(current_point, neighbor_point):
                    if current_point not in graph:
                        graph[current_point] = []
                    graph[current_point].append(neighbor_point)

    return graph

def all_paths_dfs(start, end, path=None, current_min_steepness = float('inf'), max_steepness = 0):
    if path is None:
        path = []
    path = path + [start]

    if start == end:
        current_min_steepness = min(current_min_steepness, max_steepness)
        return [path], current_min_steepness

    if start not in graph:
        return [], current_min_steepness

    paths = []
    for neighbor in graph[start]:
        if neighbor not in path:
            x_1, y_1 = start
            x_2, y_2 = neighbor
            new_steepness = abs(original_plot_sheet.iloc[x_1, y_1] - original_plot_sheet.iloc[x_2, y_2])
            new_paths, current_min_steepness = all_paths_dfs(neighbor, end, path, current_min_steepness, max(max_steepness, new_steepness))
            for p in new_paths:
                paths.append(p)

    return paths, current_min_steepness

def all_paths_bfs(start, end):
    queue = deque([(start, [start], 0)])
    paths = []
    current_min_steepness = float('inf')

    while queue:
        current, path, max_steepness = queue.popleft()

        if current == end:
            paths.append(path)
            current_min_steepness = min(current_min_steepness, max_steepness)

        if current in graph:
            for neighbor in graph[current]:
                if neighbor not in path:
                    x_1, y_1 = current
                    x_2, y_2 = neighbor
                    new_steepness = abs(original_plot_sheet.iloc[x_1, y_1] - original_plot_sheet.iloc[x_2, y_2])
                    queue.append((neighbor, path + [neighbor], max(max_steepness, new_steepness)))

    return current_min_steepness

def calculate_steepness_matrix(original_plot_sheet):
    X, Y = original_plot_sheet.shape
    steepness_matrix = np.zeros((X, Y, X, Y))
    num = 1

    for (i, j), (k, l) in itertools.product(np.ndindex(X, Y), repeat=2):
        if i <= k:
            paths, steepness_matrix[i, j, k, l] = all_paths_dfs((i, j), (k, l), path=None)
        print(f"({i},{j}),({k},{l}) , {num}/{X * Y * X * Y}")
        num = num +1

    steepness_matrix = np.maximum(steepness_matrix, steepness_matrix.transpose((2, 3, 0, 1)))

    return steepness_matrix

def generate_excel_with_border(original_plot_sheet, output_folder, output_filename):
    temp_excel = "temp.xlsx"
    original_plot_sheet.to_excel(temp_excel, index=False)

    X, Y = original_plot_sheet.shape

    wb = load_workbook(temp_excel)
    ws = wb.active

    ws.insert_rows(1)
    ws.insert_cols(1)

    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:

            cell.alignment = Alignment(horizontal='center', vertical='center')

            row_idx, col_idx = cell.row, cell.column
            print(row_idx-2, col_idx-2)
            if row_idx-2 < X and col_idx-2 < Y:
                point_index = A.index((row_idx - 2, col_idx - 2))
                print(point_index)

                border = Border(top=Side(), bottom=Side(), left=Side(), right=Side())

                if (point_index - Y) >= 0 and labels[point_index - Y] != labels[point_index]:
                    border.top = Side(style='medium', color='FF0000')
                if (point_index + Y) < X * Y and labels[point_index + Y] != labels[point_index]:
                    border.bottom = Side(style='medium', color='FF0000')
                if (point_index % Y) != 0 and labels[point_index - 1] != labels[point_index]:
                    border.left = Side(style='medium', color='FF0000')
                if ((point_index + 1) % Y) != 0 and labels[point_index + 1] != labels[point_index]:
                    border.right = Side(style='medium', color='FF0000')

                if  point_index >= 0 and point_index < Y:
                    border.top = Side(style='medium', color='FF0000')
                if  point_index >= X*Y-Y and point_index < X*Y:
                    border.bottom = Side(style='medium', color='FF0000')
                if (point_index % Y) == 0:
                    border.left = Side(style='medium', color='FF0000')
                if ((point_index + 1) % Y) == 0:
                    border.right = Side(style='medium', color='FF0000')

                below_cell_coords = (row_idx + 1, col_idx)

                below_cell = ws.cell(row=below_cell_coords[0], column=below_cell_coords[1])
                below_cell.border = border

    os.makedirs(output_folder, exist_ok=True)
    output_excel_path = os.path.join(output_folder, output_filename)

    wb.save(output_excel_path)
    wb.close()
    os.remove(temp_excel)

output_path_1 = ""
output_folder_1 = ""
output_filename_1 = ''

original_plot_sheet = pd.read_excel("", header=None)

point1 = (1, 1)
point2 = (1, 2)

A = []
for i in range(original_plot_sheet.shape[0]):
    for j in range(original_plot_sheet.shape[1]):
        A.append((i, j))

graph = adjacency_list(original_plot_sheet)

start_time = time.time()
steepness_matrix = calculate_steepness_matrix(original_plot_sheet)
end_time= time.time()
elapsed_time = end_time - start_time
print(f"Elapsed time: {elapsed_time} seconds")

start_time = time.time()
dbscan = DBSCAN(eps=1.5, min_samples=4)
labels = dbscan.fit_predict(A)
end_time= time.time()
elapsed_time = end_time - start_time
print(f"Elapsed time: {elapsed_time} seconds")

generate_excel_with_border(original_plot_sheet, output_folder_1, output_filename_1)